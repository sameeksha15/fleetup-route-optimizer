"""Parse a company's own CSV / Excel order file into validated order rows.

Companies keep their orders in spreadsheets with their own column names, so the
parser is deliberately tolerant: headers are matched case-insensitively against a
set of synonyms, addresses without coordinates are geocoded, and each row is
validated independently — one bad row never sinks the whole file. The result is a
per-row preview (ok / error) the user confirms before anything is written.
"""

from __future__ import annotations

import csv
import io
from collections.abc import Callable
from typing import Any

# Canonical field -> accepted header spellings (compared after normalization).
_SYNONYMS: dict[str, set[str]] = {
    "reference": {"reference", "ref", "order_id", "orderid", "order", "order_no", "orderno", "invoice", "awb"},
    "recipient": {"recipient", "customer", "customer_name", "name", "consignee", "receiver", "client"},
    "address": {"address", "delivery_address", "addr", "location", "destination", "shipping_address"},
    "latitude": {"latitude", "lat", "y"},
    "longitude": {"longitude", "lng", "lon", "long", "x"},
    "weight_kg": {"weight", "weight_kg", "wt", "kg", "weightkg"},
    "length_cm": {"length", "length_cm", "l", "len"},
    "width_cm": {"width", "width_cm", "w", "breadth", "b"},
    "height_cm": {"height", "height_cm", "h", "ht"},
    "volume_m3": {"volume", "volume_m3", "vol", "cbm", "m3"},
    "priority": {"priority", "prio", "importance"},
    "window_start": {"window_start", "window_start_min", "start", "from", "window_from", "tw_start", "slot_start", "start_time"},
    "window_end": {"window_end", "window_end_min", "end", "to", "window_to", "tw_end", "slot_end", "end_time"},
    "warehouse": {"warehouse", "warehouse_name", "warehouse_id", "depot", "source", "origin", "hub"},
}
_HEADER_TO_FIELD = {spelling: field for field, spellings in _SYNONYMS.items() for spelling in spellings}

_PRIORITY_WORDS = {"low": 0, "l": 0, "medium": 1, "med": 1, "m": 1, "normal": 1, "high": 2, "h": 2, "urgent": 2}

# The canonical template offered for download (order matters for the example).
TEMPLATE_HEADERS = [
    "reference", "recipient", "address", "latitude", "longitude", "weight_kg",
    "length_cm", "width_cm", "height_cm", "volume_m3", "priority",
    "window_start", "window_end", "warehouse",
]
TEMPLATE_EXAMPLE = [
    "INV-1001", "Sharma Traders", "Andheri East, Mumbai", "", "", "25",
    "60", "40", "30", "", "high", "09:00", "13:00", "Andheri East DC",
]


def _norm_header(raw: str) -> str:
    out = []
    for ch in raw.strip().lower():
        out.append(ch if ch.isalnum() else "_")
    return "_".join(part for part in "".join(out).split("_") if part)


def _read_table(file_bytes: bytes, filename: str) -> list[dict[str, str]]:
    """Read CSV or XLSX bytes into a list of {canonical_field: raw_string} rows."""
    name = filename.lower()
    if name.endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            return []
        fields = [_HEADER_TO_FIELD.get(_norm_header(str(h)) if h is not None else "") for h in header]
        table: list[dict[str, str]] = []
        for row in rows_iter:
            record = {}
            for field, value in zip(fields, row):
                if field and value is not None:
                    record[field] = str(value).strip()
            if any(record.values()):
                table.append(record)
        wb.close()
        return table

    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    try:
        header = next(reader)
    except StopIteration:
        return []
    fields = [_HEADER_TO_FIELD.get(_norm_header(h)) for h in header]
    table = []
    for row in reader:
        record = {}
        for field, value in zip(fields, row):
            if field and value is not None and str(value).strip():
                record[field] = str(value).strip()
        if any(record.values()):
            table.append(record)
    return table


def _parse_float(value: str | None) -> tuple[float | None, bool]:
    """(number or None, ok). Blank is ok; unparseable is not."""
    if value is None or value.strip() == "":
        return None, True
    try:
        return float(value), True
    except ValueError:
        return None, False


def _parse_priority(value: str | None) -> tuple[int, bool]:
    if value is None or value.strip() == "":
        return 1, True
    v = value.strip().lower()
    if v in _PRIORITY_WORDS:
        return _PRIORITY_WORDS[v], True
    try:
        n = int(float(v))
    except ValueError:
        return 1, False
    return (n, True) if 0 <= n <= 2 else (1, False)


def _parse_time(value: str | None) -> tuple[int | None, bool]:
    """Accept 'HH:MM', 'H:MM AM/PM', or integer minutes-since-midnight."""
    if value is None or value.strip() == "":
        return None, True
    raw = value.strip()
    # Plain integer -> minutes since midnight.
    try:
        minutes = int(float(raw))
        return (minutes, True) if 0 <= minutes <= 1439 else (None, False)
    except ValueError:
        pass
    text = raw.upper().replace(".", ":")
    meridiem = None
    for tag in ("AM", "PM"):
        if tag in text:
            meridiem = tag
            text = text.replace(tag, "").strip()
    if ":" not in text:
        return None, False
    try:
        hh, mm = (int(p) for p in text.split(":")[:2])
    except ValueError:
        return None, False
    if meridiem == "PM" and hh != 12:
        hh += 12
    if meridiem == "AM" and hh == 12:
        hh = 0
    if not (0 <= hh <= 23 and 0 <= mm <= 59):
        return None, False
    return hh * 60 + mm, True


def parse_orders(
    file_bytes: bytes,
    filename: str,
    *,
    warehouses: list[tuple[int, str]],
    resolve_coords: Callable[[str], tuple[float, float] | None],
) -> list[dict[str, Any]]:
    """Return one preview record per data row: {row, status, errors, order, note}."""
    wh_by_name = {name.strip().lower(): wid for wid, name in warehouses}
    wh_ids = {wid for wid, _ in warehouses}

    records = _read_table(file_bytes, filename)
    preview: list[dict[str, Any]] = []

    for i, rec in enumerate(records, start=2):  # row 1 is the header
        errors: list[str] = []
        note = ""

        reference = rec.get("reference") or None
        recipient = rec.get("recipient") or None
        address = rec.get("address") or None

        lat, lat_ok = _parse_float(rec.get("latitude"))
        lon, lon_ok = _parse_float(rec.get("longitude"))
        if not lat_ok or not lon_ok:
            errors.append("Latitude/longitude is not a number.")

        # Resolve coordinates: prefer given lat/lng, else geocode the address.
        if lat is not None and lon is not None:
            if not (-90 <= lat <= 90 and -180 <= lon <= 180):
                errors.append("Latitude/longitude out of range.")
        elif address:
            hit = resolve_coords(address)
            if hit is None:
                errors.append("Could not geocode the address; add latitude/longitude.")
            else:
                lat, lon = hit
                note = "geocoded"
        else:
            errors.append("Provide an address, or latitude and longitude.")

        if not address and lat is not None and lon is not None:
            address = f"{lat:.5f}, {lon:.5f}"

        weight, weight_ok = _parse_float(rec.get("weight_kg"))
        if not weight_ok:
            errors.append("Weight is not a number.")
        elif weight is None or weight <= 0:
            errors.append("Weight (kg) is required and must be positive.")

        dims = {}
        for key in ("length_cm", "width_cm", "height_cm"):
            val, ok = _parse_float(rec.get(key))
            if not ok:
                errors.append(f"{key.split('_')[0].title()} is not a number.")
            elif val is not None and val <= 0:
                errors.append(f"{key.split('_')[0].title()} must be positive.")
            dims[key] = val
        has_all_dims = all(dims[k] is not None for k in dims)
        has_any_dims = any(dims[k] is not None for k in dims)
        if has_any_dims and not has_all_dims:
            errors.append("Provide all three dimensions (length, width, height) or none.")

        volume, vol_ok = _parse_float(rec.get("volume_m3"))
        if not vol_ok:
            errors.append("Volume is not a number.")
        if volume is None and has_all_dims:
            volume = round(dims["length_cm"] * dims["width_cm"] * dims["height_cm"] / 1_000_000, 3)
        if volume is None and not has_all_dims:
            errors.append("Provide volume or all three dimensions.")
        elif volume is not None and volume <= 0:
            errors.append("Volume must be positive.")

        priority, prio_ok = _parse_priority(rec.get("priority"))
        if not prio_ok:
            errors.append("Priority must be low/medium/high or 0-2.")

        ws, ws_ok = _parse_time(rec.get("window_start"))
        we, we_ok = _parse_time(rec.get("window_end"))
        if not ws_ok or not we_ok:
            errors.append("Delivery window time is invalid (use HH:MM).")
        elif (ws is None) != (we is None):
            errors.append("Provide both window start and end, or neither.")
        elif ws is not None and we is not None and ws >= we:
            errors.append("Window start must be before end.")

        warehouse_id = None
        wh_raw = rec.get("warehouse")
        if wh_raw:
            key = wh_raw.strip().lower()
            if key in wh_by_name:
                warehouse_id = wh_by_name[key]
            elif wh_raw.isdigit() and int(wh_raw) in wh_ids:
                warehouse_id = int(wh_raw)
            else:
                errors.append(f"Warehouse '{wh_raw}' is not one of your warehouses.")

        order = {
            "reference": reference,
            "recipient": recipient,
            "address": address,
            "latitude": lat,
            "longitude": lon,
            "weight_kg": weight,
            "volume_m3": volume,
            "length_cm": dims["length_cm"],
            "width_cm": dims["width_cm"],
            "height_cm": dims["height_cm"],
            "priority": priority,
            "window_start_min": ws,
            "window_end_min": we,
            "warehouse_id": warehouse_id,
        }
        preview.append(
            {
                "row": i,
                "status": "ok" if not errors else "error",
                "errors": errors,
                "note": note,
                "order": order,
            }
        )

    return preview
